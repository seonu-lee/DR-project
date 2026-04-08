"""
서울시 에너지 사용량 엑셀 합치기
- energy_2018.xlsx ~ energy_2025.xlsx → seoul_energy_total.xlsx
"""

import pandas as pd
import glob
import re
from openpyxl.styles import Font, PatternFill, Alignment

# ── 경로 설정 ──────────────────────────────────────────────────────────────────
INPUT_DIR   = r"C:\Users\seonu\Documents\DR-project\data"   # 엑셀 파일들이 있는 폴더 경로
OUTPUT_FILE = r"C:\Users\seonu\Documents\DR-project\data\seoul_energy_total.xlsx"  # 저장할 파일 경로
# ──────────────────────────────────────────────────────────────────────────────

import os
all_dfs = []

for filepath in sorted(glob.glob(os.path.join(INPUT_DIR, "energy_20*.xlsx"))):
    year = int(re.search(r"(\d{4})", filepath).group(1))

    # 헤더가 2행(1행=제목, 2행=컬럼명)이므로 header=1로 읽기
    df = pd.read_excel(filepath, header=1)

    # 실제 컬럼명 정규화
    df.columns = [str(c).strip() for c in df.columns]

    # 번호/자치구명/행정동명/계/1월~12월 컬럼만 추출
    keep = ["번호", "자치구명", "행정동명", "계"] + [f"{i}월" for i in range(1, 13)]
    # 컬럼명이 약간 다를 수 있으므로 유연하게 매핑
    col_map = {}
    for c in df.columns:
        c_clean = c.replace(" ", "")
        if c_clean == "번호":             col_map[c] = "번호"
        elif "자치구" in c_clean:         col_map[c] = "자치구명"
        elif "행정동" in c_clean:         col_map[c] = "행정동명"
        elif c_clean == "계":             col_map[c] = "계"
        else:
            m = re.match(r"(\d+)월", c_clean)
            if m:
                col_map[c] = f"{m.group(1)}월"

    df = df.rename(columns=col_map)

    # 필요한 컬럼만 선택 (없는 월은 None으로)
    result_cols = ["번호", "자치구명", "행정동명", "계"] + [f"{i}월" for i in range(1, 13)]
    for col in result_cols:
        if col not in df.columns:
            df[col] = None

    df = df[result_cols].copy()

    # 번호가 숫자인 행만 (헤더/합계 행 제거)
    df = df[pd.to_numeric(df["번호"], errors="coerce").notna()].copy()
    df["번호"] = pd.to_numeric(df["번호"]).astype(int)

    # 연도 컬럼 추가
    df.insert(1, "연도", year)

    all_dfs.append(df)
    print(f"  {filepath}: {len(df)}행 로드")

total_df = pd.concat(all_dfs, ignore_index=True)

# 번호 재부여
total_df["번호"] = range(1, len(total_df) + 1)

# 저장
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    total_df.to_excel(writer, index=False, sheet_name="전체")
    ws = writer.sheets["전체"]

    fill = PatternFill("solid", start_color="4472C4", end_color="4472C4")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    col_widths = {"A": 7, "B": 8, "C": 12, "D": 14, "E": 14}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w
    for col in "FGHIJKLMNOPQ":
        ws.column_dimensions[col].width = 12

print(f"\n완료: {OUTPUT_FILE} ({len(total_df):,}행)")